import os
import openpyxl
import pandas as pd
import math
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QApplication, QDialog, QMessageBox
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from QACO import Ui_dialog
import sys
mpl.use('Qt5Agg')

class Ui_ACOdialog(QDialog):
    def __init__(self, dic_Str, dic_CC):
        super().__init__()
        self.ui = Ui_dialog()
        self.ui.setupUi(self)
        self.dic_Str = self.Sortdic_Str(dic_Str)  # 条带信息字典
        self.dic_CC = dic_CC  # 缆机信息字典
        self.fig, self.ax = plt.subplots(figsize=(10,5))
        self.canvas = FigureCanvasQTAgg(self.fig)
        mpl.rc("font", family='Times New Roman', weight="bold")

    @pyqtSlot()
    def on_act_CalculateBut_clicked(self):
        self.num_Str = len(self.dic_Str)   # 条带数量
        self.num_CC = len(self.dic_CC)   # 缆机数量
        self.CC_Str = self.initializeCC_Str()  # 缆机-条带矩阵模型
        self.NumAnts = int(self.ui.LE_NumAnts.text())   # 蚂蚁数量
        self.Alpha = float(self.ui.LE_Alpha.text())   # 信息素重要程度因子
        self.Beta = float(self.ui.LE_Beta.text())  # 启发函数重要因子
        self.Rho = float(self.ui.LE_Rho.text())  # 信息素持久因子
        self.CoeforAlpha = float(self.ui.LE_CoeforAlpha.text())   # 信息素常量系数
        self.CoeforBeta = float(self.ui.LE_CoeforBeta.text())   # 启发函数常量系数
        self.InitAlpha = float(self.ui.LE_InitAlpha.text())   # 初始信息素
        self.IterMax = int(self.ui.LE_IterMax.text())   # 迭代次数
        self.VRack = float(self.ui.LE_VRack.text())   # 缆机机架运行速度
        self.VTCage = float(self.ui.LE_VTCage.text())   # 满载水平运行速度
        self.VReCage = float(self.ui.LE_VReCage.text())   # 空载返回运行速度
        self.VDownCage = float(self.ui.LE_VDownCage.text())    # 满载垂直下降速度
        self.VUpCage = float(self.ui.LE_VUpCage.text())   # 空载上升运行速度
        self.CageV = float(self.ui.LE_CageV.text())   # 吊罐体积
        self.TimeLoad = int(self.ui.LE_TimeLoad.text())   # 装料时间90s
        self.TimeUnload = int(self.ui.LE_TimeUnload.text())    # 卸料时间100s
        self.SafeDis = float(self.ui.LE_SafeDis.text())    # 缆机之间的安全距离
        self.Tau = [[self.InitAlpha for _ in range(self.num_CC)] for _ in range(self.num_Str)]  # 初始化信息素矩阵，一个num_Band行，num_Cable列的矩阵
        self.Table = [[0 for _ in range(self.num_Str)] for _ in range(self.NumAnts)]  # 生成的蚁群，给每个蚂蚁赋初始值

        self.ACO()

    def initializeCC_Str(self):
        CC_Str = np.zeros((self.num_Str, self.num_CC, 2), dtype="int")
        for row in range(self.num_Str):
            for col in range(self.num_CC):
                CC_Str[row][col] = [row+1, col+1]
        return CC_Str

    def Sortdic_Str(self,dic_Str):
        dic_StrTem = {}
        StrTem = sorted(dic_Str.items(), key=lambda x: x[1][1])
        num = 1
        for obj in StrTem:
            dic_StrTem[num] = [obj[0], obj[1][:8]]
            num = num+1
        return dic_StrTem

    def ACO(self):
        self.best_T = math.inf  # 最好的方案，工期
        self.best_Path = None   # 最好的方案，任务分配关系
        self.best_TStrlist = None   # 最好的方案，条带浇筑完成时间节点
        self.best_TCC_Work = None  # 最好的方案，缆机实际工作时长
        self.best_TCC_Finish = None  # 最好的方案，缆机完成任务时间节点
        self.best_T2 = math.inf  # 第二好的方案，工期
        self.best_Path2 = None   # 第二好的方案，任务分配关系
        self.best_TStrlist2 = None   # 第二好的方案，条带浇筑完成时间节点
        self.best_TCC_Work2 = None  # 第二好的方案，缆机实际工作时长
        self.best_TCC_Finish2 = None  # 第二好的方案，缆机完成任务时间节点
        self.best_T3 = math.inf  # 第三好的方案，工期
        self.best_Path3 = None   # 第三好的方案，任务分配关系
        self.best_TStrlist3 = None   # 第三好的方案，条带浇筑完成时间节点
        self.best_TCC_Work3 = None  # 第三好的方案，缆机实际工作时长
        self.best_TCC_Finish3 = None  # 第三好的方案，缆机完成任务时间节点

        result = []
        for iter in range(self.IterMax):
            # 生成新的蚁群
            #  TCC_Work_ants用来记录每个解中每个缆机的实际工作时间时间；T_ants记录所有蚂蚁的路径；
            #  TStr_list记录每个条带被完成浇筑的时间节点；TPour_CCDic记录每个缆机的完工时间。
            T_ants, TCC_Work_ants, TStr_list, TCC_Finish = self.ConstructSolution()
            temSet_T_ants = list(set(T_ants.copy()))
            temList_T_ants = sorted(temSet_T_ants)
            if len(temSet_T_ants) >= 3:
                for obj in temList_T_ants[:3]:
                    Num_index = T_ants.index(obj)
                    if obj<self.best_T:
                        self.best_T = obj
                        self.best_Path = self.Table[Num_index]
                        self.best_TStrlist = TStr_list[Num_index]
                        self.best_TCC_Work = TCC_Work_ants[Num_index]
                        self.best_TCC_Finish = TCC_Finish[Num_index]
                    elif obj<self.best_T2:
                        self.best_T2 = obj
                        self.best_Path2 = self.Table[Num_index]
                        self.best_TStrlist2 = TStr_list[Num_index]
                        self.best_TCC_Work2 = TCC_Work_ants[Num_index]
                        self.best_TCC_Finish2 = TCC_Finish[Num_index]
                    elif obj<self.best_T3:
                        self.best_T3 = obj
                        self.best_Path3 = self.Table[Num_index]
                        self.best_TStrlist3 = TStr_list[Num_index]
                        self.best_TCC_Work3 = TCC_Work_ants[Num_index]
                        self.best_TCC_Finish3 = TCC_Finish[Num_index]
            else:
                for obj in temList_T_ants:
                    Num_index = T_ants.index(obj)
                    if obj < self.best_T:
                        self.best_T = obj
                        self.best_Path = self.Table[Num_index]
                        self.best_TStrlist = TStr_list[Num_index]
                        self.best_TCC_Work = TCC_Work_ants[Num_index]
                        self.best_TCC_Finish = TCC_Finish[Num_index]
                    elif obj < self.best_T2:
                        self.best_T2 = obj
                        self.best_Path2 = self.Table[Num_index]
                        self.best_TStrlist2 = TStr_list[Num_index]
                        self.best_TCC_Work2 = TCC_Work_ants[Num_index]
                        self.best_TCC_Finish2 = TCC_Finish[Num_index]
                    elif obj < self.best_T3:
                        self.best_T3 = obj
                        self.best_Path3 = self.Table[Num_index]
                        self.best_TStrlist3 = TStr_list[Num_index]
                        self.best_TCC_Work3 = TCC_Work_ants[Num_index]
                        self.best_TCC_Finish3 = TCC_Finish[Num_index]

            # 更新信息素
            self.update_Tau(T_ants)

            # 绘制迭代过程
            rlist = []
            for Slist in TCC_Work_ants:
                maxT = max(Slist)
                stdlist = []
                for s in Slist:
                    stdlist.append(s/maxT)
                rlist.append(np.std(stdlist))

            slist = []
            for i in range(len(T_ants)):
                v1 = T_ants[i]/3000
                v2 = rlist[i]
                s = v1
                slist.append(s)
            result.append(min(slist))
            self.ax.plot(result, color = 'r')
            self.ax.set_xlabel("Iterations", fontsize=20, color = 'k')
            self.ax.set_ylabel("Objective function value", fontsize=20, color = 'k')
            plt.suptitle("缆机站位优化迭代计算过程", fontsize=12, color='red')
            plt.pause(0.1)
        plt.close()
        self.layout = self.ui.ACOLayout
        self.layout.addWidget(self.canvas)

    @pyqtSlot()
    def on_act_Save_clicked(self):
        desktop_path = ".\Alternative\\"
        # 文件1
        full_path1 = desktop_path + "Plan1" + '.xlsx'
        if os.path.exists(full_path1):
            os.remove(full_path1)
        excel1 = openpyxl.Workbook()
        sht1 = excel1.create_sheet(index=0,title="data")
        sht1['A1'] = self.best_T
        for col in range(len(self.best_TCC_Work)):
            absc = self.best_TCC_Work[col]
            sht1.cell(row = 2, column = col+1, value = str(absc))
        for col in range(len(self.best_TCC_Finish)):
            sht1.cell(row = 3, column = col+1, value = self.best_TCC_Finish[col+1])
        nowDam_path = self.path_damwbs(self.best_Path)
        # 暂时
        value_list = []
        for value in self.best_TStrlist.values():
            value_list.append(value)

        for row in range(len(nowDam_path)):
            sht1.cell(row =4 + row, column =1, value = str(nowDam_path[row][0]))
            sht1.cell(row =4 + row, column =2, value = str(nowDam_path[row][1]))
            # 暂时
            sht1.cell(row =4 + row, column =3, value = str(value_list[row]))

        excel1.save(full_path1)

        # 文件2
        full_path2 = desktop_path + "Plan2" + '.xlsx'
        if os.path.exists(full_path2):
            os.remove(full_path2)
        excel2 = openpyxl.Workbook()
        sht2 = excel2.create_sheet(index=0,title="data")
        sht2['A1'] = self.best_T2
        for col in range(len(self.best_TCC_Work2)):
            sht2.cell(row =2, column =col+1, value = self.best_TCC_Work2[col])
        for col in range(len(self.best_TCC_Finish2)):
            sht2.cell(row =3, column =col+1, value  = self.best_TCC_Finish2[col+1])
        nowDam_path = self.path_damwbs(self.best_Path2)
        for row in range(len(nowDam_path)):
            sht2.cell(row =4 + row, column =1, value = str(nowDam_path[row][0]))
            sht2.cell(row =4 + row, column =2, value = str(nowDam_path[row][1]))
        excel2.save(full_path2)

        # 文件3
        full_path3 = desktop_path + "Plan3" + '.xlsx'
        if os.path.exists(full_path3):
            os.remove(full_path3)
        excel3 = openpyxl.Workbook()
        sht3 = excel3.create_sheet(index=0,title="data")
        sht3['A1'] = self.best_T3
        for col in range(len(self.best_TCC_Work3)):
            sht3.cell(row =2, column =col+1, value = self.best_TCC_Work3[col])
        for col in range(len(self.best_TCC_Finish3)):
            sht3.cell(row =3, column =col+1, value = self.best_TCC_Finish3[col+1])
        nowDam_path = self.path_damwbs(self.best_Path3)
        for row in range(len(nowDam_path)):
            sht3.cell(row =4 + row, column =1, value = str(nowDam_path[row][0]))
            sht3.cell(row =4 + row, column =2, value = str(nowDam_path[row][1]))
        excel3.save(full_path3)

        dleTitle = "消息提示"
        strInfo = "保存完毕！"
        QMessageBox.information(self, dleTitle, strInfo)

    def path_damwbs(self,paths):
        nowDam_path = []
        for path in paths:
            tem = self.dic_Str[path[0]][0]
            nowDam_path.append([tem,path[1]])
        return nowDam_path
            
    def update_Tau(self, T_ants):
        print(self.Tau)
        for i in range(len(self.Tau)):
            for j in range(len(self.Tau[i])):
                self.Tau[i][j] = self.Rho * self.Tau[i][j]

        for num in range(len(T_ants)):
            if T_ants[num] == min(T_ants):
                unit_Tau = self.CoeforAlpha/T_ants[num]
                for i in range(len(self.Table[num])):
                    row = self.Table[num][i][0]
                    col = self.Table[num][i][1]
                    self.Tau[row-1][col-1] = self.Tau[row-1][col-1] + unit_Tau

    def ConstructSolution(self):
        T_ants = []
        TStr_list = []
        TCC_Work_ants = []
        TCC_Finish = []
        for ant in range(self.NumAnts):
            Path_ant = {}  # 记录单个蚂蚁的路径及浇筑每个条带的浇筑时间
            dic_CC_Recording = self.dic_CC.copy()  # 每次开始前都重置缆机位置
            TWork_CC = [0.0 for _ in range(self.num_CC)]  # 用来记录每个缆机的实际工作时间时间
            Limit_CC = []  # 用来记录每个缆机覆盖区域的上下界限
            for i in range(self.num_CC):
                Limit_CC.append([0.0, math.inf])
            startNode = self.CC_Str[0][0]  # 以网格左上角节点[1,1]开始
            self.Table[ant][0] = [startNode[0], startNode[1]]
            current_col = 0  # 表示当前所在列，也即缆机
            # 计算浇筑时间
            T_Cycle, Eta = self.compute_tau(0, current_col, dic_CC_Recording)
            # 记录单个蚂蚁的路径及浇筑每个条带的浇筑时间
            Path_ant[(startNode[0], startNode[1])] = T_Cycle
            # 更新每个缆机的浇筑时间
            TWork_CC[current_col] = float(TWork_CC[current_col]) + float(T_Cycle)
            # 更新每个缆机的上下界限
            uplimit_y = self.dic_Str[startNode[0]][1][5]
            downlimit_y = self.dic_Str[startNode[0]][1][6]
            Limit_CC[current_col] = [uplimit_y, downlimit_y]
            # 更新缆机位置
            dic_CC_Recording[current_col+1] = [dic_CC_Recording[current_col+1][0], self.dic_Str[1][1][1], dic_CC_Recording[current_col+1][2]]
            for row in range(self.num_Str-1):
                # 判断缆机是不是处于最后一个，即是不是最后一列
                if current_col < self.num_CC - 1:
                # 计算下一行竖直和斜向节点的信息素值和启发信息值
                    t_Vertical, Eta_Vertica = self.compute_tau(row+1, current_col, dic_CC_Recording)
                    t_Diagonal, Eta_Diagonal = self.compute_tau(row+1, current_col+1, dic_CC_Recording)
                    P_Vertical_tem = self.Tau[row+1][current_col] ** self.Alpha * Eta_Vertica ** self.Beta
                    P_Diagonal_tem = self.Tau[row+1][current_col+1] ** self.Alpha * Eta_Diagonal ** self.Beta
                    P_Vertical = P_Vertical_tem/(P_Vertical_tem+P_Diagonal_tem)
                    P_Diagonal = P_Diagonal_tem/(P_Vertical_tem+P_Diagonal_tem)
                    current_col, T_Cycle = self.node_choose(P_Vertical, P_Diagonal, Limit_CC, current_col, t_Vertical, t_Diagonal)
                else:
                    T_Cycle, Eta_Vertica = self.compute_tau(row + 1, current_col, dic_CC_Recording)
                row = row + 1
                # 更新路径
                self.Table[ant][row] = [row+1, current_col+1]
                # 记录单个蚂蚁的路径及浇筑每个条带的浇筑时间
                Path_ant[(row+1, current_col+1)] = T_Cycle
                # 更新每个缆机的浇筑时间
                TWork_CC[current_col] = float(TWork_CC[current_col]) + float(T_Cycle)
                # 更新每个缆机覆盖区域的上下界限
                location_tem = self.CC_Str[row][current_col]
                uplimit_y = self.dic_Str[location_tem[0]][1][5]
                downlimit_y = self.dic_Str[location_tem[0]][1][6]
                if uplimit_y > Limit_CC[current_col][0]:
                    Limit_CC[current_col][0] = uplimit_y
                if downlimit_y < Limit_CC[current_col][1]:
                    Limit_CC[current_col][1] = downlimit_y
                # 更新缆机位置
                dic_CC_Recording[current_col + 1] = [dic_CC_Recording[current_col + 1][0],
                                                             self.dic_Str[1][1][1],
                                                             dic_CC_Recording[current_col + 1][2]]

            TPour_CCDic, T_main, CC_Str_T = self.SimulationPour(Path_ant)   # TPour_CC每个缆机的最终的浇筑周期，包括停歇。
            TStr_list.append(CC_Str_T)
            # print(CC_Str_T)
            TCC_Finish.append(TPour_CCDic)

            T_ants.append(T_main)
            TCC_Work_ants.append(TWork_CC)   # 用来记录每个缆机的实际工作时间时间

        return T_ants, TCC_Work_ants, TStr_list, TCC_Finish


    # 模拟缆机浇筑过程
    def SimulationPour(self,Path_ant):
        T_main = 0.0
        TPour_CCDic = dict.fromkeys(range(1,self.num_CC+1), 0)  # 用来记录每个缆机的浇筑时间
        CC_Sever_Str = {}  # 用来记录每个缆机浇筑的条带对象。
        CC_Str_T = {}   # 用来记录浇筑过程中，每个条带被服务的次序。
        for num in range(1, self.num_CC + 1):
            temList = []
            for key, value in Path_ant.items():
                if num == key[1]:
                    temList.append([key[0],value])
            CC_Sever_Str[num] = temList
        while True:
            T_tem = []
            for num in range(1, self.num_CC + 1):
                StrT_List = CC_Sever_Str[num]
                if StrT_List:
                    T_tem.append(StrT_List[0][1] + TPour_CCDic[num])
                else:
                    T_tem.append(math.inf)
            Move_CC = T_tem.index(min(T_tem))+1
            str_WBS0 = CC_Sever_Str[Move_CC][0][0]
            T_main = min(T_tem)
            CC_Str_T[(str_WBS0, Move_CC)] = T_main   #  (条带，缆机)=时间节点
            TPour_CCDic[Move_CC] = min(T_tem)
            CC_Sever_Str[Move_CC].pop(0)
            # 判断是否符合安全距离
            if Move_CC < self.num_CC:
                if CC_Sever_Str[Move_CC] and CC_Sever_Str[Move_CC+1]:
                    str_WBS1 = CC_Sever_Str[Move_CC][0][0]
                    str_WBS2 = CC_Sever_Str[Move_CC+1][0][0]
                    # 判断相邻条带间的距离
                    d = abs(self.dic_Str[str_WBS1][1][1]-self.dic_Str[str_WBS2][1][1])
                    if d < self.SafeDis:
                        d_time = T_tem[Move_CC]-T_tem[Move_CC-1]
                        CC_Sever_Str[Move_CC][0][1] = CC_Sever_Str[Move_CC][0][1] + d_time
            JudgeCondition = 0
            for value in CC_Sever_Str.values():
                JudgeCondition = JudgeCondition + len(value)
            if JudgeCondition == 0:
                break
        return TPour_CCDic, T_main, CC_Str_T

    # 轮盘赌选择
    def node_choose(self, P_Vertical, P_Diagonal, Limit_CC, current_col, t_Vertical, t_Diagonal):
        t = 0
        uplimit_y = Limit_CC[current_col][0]
        downlimit_y = Limit_CC[current_col][1]
        if (uplimit_y-downlimit_y) > self.SafeDis:  # 如果上下限符合安全距离, 可以进行轮盘赌
            x = np.random.rand()
            if x < P_Vertical:
                next_current_col = current_col
                t = t_Vertical
            else:
                next_current_col = current_col+1
                t = t_Diagonal
        else:
            next_current_col = current_col
            t = t_Vertical
        return next_current_col, t

    def compute_tau(self, row, col, dic_CC_Recording):
        str = self.CC_Str[row][col][0]
        CC = self.CC_Str[row][col][1]
        Str_data = self.dic_Str[str][1]
        Str_x = float(Str_data[0])
        Str_y = float(Str_data[1])
        Str_z = float(Str_data[2])
        Str_V = float(Str_data[3])
        CC_xyz = dic_CC_Recording[CC]
        CC_x = CC_xyz[0]
        CC_y = CC_xyz[1]
        CC_z = CC_xyz[2]
        num = math.ceil(Str_V / self.CageV)  # 一个条带需要的吊罐数量
        T_transport = abs(CC_y - Str_y) / self.VRack + abs(CC_x - Str_x) / self.VTCage + \
            abs(CC_z - Str_z - 20) / self.VDownCage + abs(CC_z - Str_z - 20) / self.VUpCage + abs(CC_x - Str_x) / self.VReCage
        T_Cycle = num*(T_transport + self.TimeLoad + self.TimeUnload)/60  # 除以60换算为以分钟为计时单位
        # 计算启发信息值
        Eta = self.CoeforBeta / T_Cycle
        return T_Cycle, Eta




if __name__== "__main__":
    app = QApplication(sys.argv)
    form = Ui_ACOdialog()
    form.show()
    sys.exit(app.exec_())