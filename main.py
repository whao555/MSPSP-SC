import os
import re
import sys
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import matplotlib as mpl
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QMainWindow, QApplication, QDialog, QTableWidgetItem, QAbstractItemView, QSplitter
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, NavigationToolbar2QT as NavigationToolbar
from QMainWindow import Ui_MainWindow
from PourEleData import QPourEleInputForm
from ACOWidget import Ui_ACOdialog
from BuildDam3D import BulidDam3D

class myMainWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.DamSecEle = self.initializeDamEle()
        self.dic_CC = {}
        self.plan = {}
        self.selectedObj = None

        mpl.rcParams['font.sans-serif'] = ['kaiTi', 'SimHei']   # 显示汉字
        mpl.rcParams['font.size'] = 16
        mpl.rcParams['axes.unicode_minus'] = False
        self.colorlist = ['red', 'yellow', 'green', 'cyan', 'blue', 'purple','orange']  # 颜色列表
        self.ui.tableWidget_2D.itemClicked.connect(self.show_data)
        self.__iniFigure()

        self.ui.tableWidget.setAlternatingRowColors(True)
        self.__tableInitialized = False

    def __iniFigure(self):
        #  添加3D页面
        self.mayavi_widget = BulidDam3D(self.DamSecEle)
        self.mayavi_ui = self.mayavi_widget.edit_traits(parent=self, kind='subpanel').control
        self.Layout = self.ui.Layout3D
        self.Layout.addWidget(self.mayavi_ui)
        #  添加2D页面
        self.__fig = mpl.figure.Figure()
        self.__fig.suptitle("缆机站位方案")
        figCanvas = FigureCanvas(self.__fig)
        #naviToobar = NavigationToolbar(figCanvas,self)
        #naviToobar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        self.LayoutFig = self.ui.Layout2D
        self.LayoutFig.addWidget(figCanvas)


    def initializeDamEle(self):
        self.poureleinput = QPourEleInputForm()
        DamSecEle = self.poureleinput.initialize()
        return DamSecEle
    # 查看当前浇筑面貌
    @pyqtSlot()
    def on_act_PourEle_triggered(self):
        poureleinput = QPourEleInputForm()
        ret = poureleinput.exec()
        if (ret == QDialog.Accepted):
            self.DamSecEle = poureleinput.DamSecEle

    # 确认选择
    @pyqtSlot()
    def on_act_SelBlk_triggered(self):
        ObjTem = []
        ExcelSurfacePath = "..\DamBody\CATIA\ExcelSurfaceData"
        selectedObj = self.mayavi_widget.selectedObj
        for obj in selectedObj:
            FileTem = "\#" + re.sub('\D','',obj[:5]) + "DAM"
            Path = ExcelSurfacePath + FileTem + "\\" + obj[:-4] + ".xlsx"
            ObjTem.append(Path)
        self.selectedObj = ObjTem

    # 查看缆机
    @pyqtSlot()
    def on_act_CC_triggered(self):
        self.ui.tableWidget.clear()
        headerText = ["编号", "供料点X", "供料点Y", "高程", "左岸平台X", "左岸平台Y",
                      "右岸平台X", "右岸平台Y"]
        self.ui.tableWidget.setColumnCount(len(headerText))
        for i in range(len(headerText)):
            headerItem = QTableWidgetItem(headerText[i])
            font = headerItem.font()
            font.setPointSize(12)
            self.ui.tableWidget.setHorizontalHeaderItem(i, headerItem)
        workbook = openpyxl.load_workbook('./CableCrane.xlsx')
        worksheet = workbook["CC"]
        self.ui.tableWidget.setRowCount(worksheet.max_row)
        rownum = 0
        for row in worksheet.iter_rows(row_offset=1):
            CCTem = [cell.value for cell in list(row)]
            if CCTem[0]:
                print(CCTem)
                self.dic_CC[rownum+1] = CCTem
                for i in range(len(CCTem)):
                    item = QTableWidgetItem(str(CCTem[i]))
                    self.ui.tableWidget.setItem(rownum, i, item)
            rownum = rownum + 1

    # 查看条带信息
    @pyqtSlot()
    def on_act_Str_triggered(self):
        self.dic_Str = {}
        self.ui.tableWidget.clear()
        headerText = ["WBS", "Block", "Strip", "Cx", "Cy", "Elevation",
                          "Volume", "Area", "uplimit_Y","downlimit_Y", "Cable",
                      "XY", "XY", "XY", "XY", "XY", "XY", "XY", "XY", "XY", "XY"]
        self.ui.tableWidget.setColumnCount(len(headerText))
        for i in range(len(headerText)):
            headerItem = QTableWidgetItem(headerText[i])
            font = headerItem.font()
            font.setPointSize(12)
            self.ui.tableWidget.setHorizontalHeaderItem(i, headerItem)
        self.ui.tableWidget.setRowCount(50)
        rownum = 0
        for obj in self.selectedObj:
            workbook = openpyxl.load_workbook(obj)
            worksheet = workbook["StripData"]
            for row in worksheet.iter_rows(row_offset=1):
                StrTem = [cell.value for cell in list(row)[1:]]
                if StrTem[0]:
                    StrTem.insert(0, rownum + 1)
                    self.dic_Str[(StrTem[1],StrTem[2])] = StrTem[3:]
                    for i in range(len(StrTem)):
                        item = QTableWidgetItem(str(StrTem[i]))
                        self.ui.tableWidget.setItem(rownum, i, item)
                    rownum = rownum + 1

    # 开始计算按钮
    @pyqtSlot()
    def on_act_Calculate_triggered(self):
        Ui_ACO = Ui_ACOdialog(self.dic_Str,self.dic_CC)
        ret = Ui_ACO.exec()
        if (ret == QDialog.Accepted):
            print("gijugk")

    # 查看计算结果按钮
    @pyqtSlot()
    def on_act_Result_triggered(self):
        self.ui.stackedWidget.setCurrentIndex(1)
        self.ui.tableWidget_2D.clear()
        headerText = ["方案", "总时长", "条带数"]
        for num in range(1, len(self.dic_CC)+1):
            strCC = "缆机" + str(num) + "利用率%"
            headerText.append(strCC)

        self.ui.tableWidget_2D.setColumnCount(len(headerText))
        for i in range(len(headerText)):
            headerItem = QTableWidgetItem(headerText[i])
            font = headerItem.font()
            font.setPointSize(12)
            self.ui.tableWidget_2D.setHorizontalHeaderItem(i, headerItem)
        self.ui.tableWidget_2D.setRowCount(10)
        desktop_path = "./Alternative/"
        planNum = 1
        for filename in os.listdir(desktop_path):
            path = os.path.join(desktop_path, filename)
            workbook = openpyxl.load_workbook(path)
            worksheet = workbook["data"]
            max_row_num = worksheet.max_row
            duration = float(worksheet['A1'].value)/60
            num_Str = len(self.dic_Str)   #  条带数量
            temCC_Str = []
            for row in range(4,max_row_num+1):
                temCC_Str.append([worksheet.cell(row = row, column = 1).value, worksheet.cell(row = row, column = 2).value])
            temCC_r = []
            for num in range(1, len(self.dic_CC) + 1):
                r_CC = float(worksheet.cell(row = 2, column = num).value)*100/float(worksheet.cell(row = 3, column = num).value)
                temCC_r.append('%.2f' % r_CC)
            self.plan[planNum] = ['%.2f' % duration, num_Str, temCC_r, temCC_Str]
            planNum = planNum +1
        row = 0
        for key, value in self.plan.items():
            item = QTableWidgetItem(str(key))
            self.ui.tableWidget_2D.setItem(row, 0, item)
            tem = value[2]
            tem.insert(0, value[1])
            tem.insert(0, value[0])
            for i in range(len(tem)):
                item = QTableWidgetItem(str(tem[i]))
                self.ui.tableWidget_2D.setItem(row, i+1, item)
            row = row+1

    # 展示3D页面
    @pyqtSlot()
    def on_act_Appearance_triggered(self):
        self.ui.stackedWidget.setCurrentIndex(0)

    # 绘制2D方案
    def __drawFigure(self, Str_xy_dic,CC_xy_dic,Max_X,Min_X,Max_Y,Min_Y):
        self.axes3 = self.__fig.add_subplot(1, 1, 1)
        self.axes3.set_xlim((Min_X-10, Max_X+10))
        self.axes3.set_ylim((Min_Y-10, Max_Y+10))
        # 绘制缆机和平台
        for valueCC in CC_xy_dic.values():
            PLCC = plt.Polygon(xy=valueCC, color='black', alpha=0.8)
            self.axes3.add_patch(PLCC)
        # 绘制条带
        for key, values in Str_xy_dic.items():
            colorPL = self.colorlist[key-1]
            for value in values:
                PL = plt.Polygon(xy=value, color=colorPL, alpha=0.8)
                self.axes3.add_patch(PL)

    # 响应表格点击事件
    def show_data(self, Item=None):
        self.__fig.clear()
        # 如果单元格对象为空
        if Item is None:
            return
        else:
            row = Item.row()  # 获取行数
            self.ShowPlan(row)

    # 根据行信息将方案绘制数据传递
    def ShowPlan(self, row):
        # 提取缆机坐标信息
        CC_xy_dic = dict.fromkeys(range(1, len(self.dic_CC) + 1), [])
        tem_X = []
        tem_Y = []

        for key, value in self.dic_CC.items():
            tem_X.append(value[4])
            tem_X.append(value[6])
            tem_Y.append(value[5])
            tem_Y.append(value[7])
            CC_xy_dic[key] = [[value[4], value[5]], [value[6], value[7]]]
        Max_X = max(tem_X)
        Min_X = min(tem_X)
        Max_Y = max(tem_Y)
        Min_Y = min(tem_Y)
        platformLeft = [[Max_X + 2, Min_Y], [Max_X + 2, Max_Y], [Max_X - 2, Max_Y], [Max_X - 2, Min_Y]]
        platformRight = [[Min_X + 2, Min_Y], [Min_X + 2, Max_Y], [Min_X - 2, Max_Y], [Min_X - 2, Min_Y]]
        CC_xy_dic["platformLeft"] = platformLeft
        CC_xy_dic["platformRight"] = platformRight
        Str_xy_dic = self.PlanDataDeal(row)
        print(Str_xy_dic)
        self.__drawFigure(Str_xy_dic, CC_xy_dic, Max_X, Min_X, Max_Y, Min_Y)
        self.__fig.canvas.draw()

    # 处理条带信息
    def PlanDataDeal(self,row):
        Str_xy_dic = dict.fromkeys(range(1, len(self.dic_CC)+1), [])
        Str_xy_list = [[] for _ in range(len(self.dic_CC))]
        pathList = self.plan[int(row)+1][3]
        for node in pathList:
            CC = int(node[1])
            dam = int(re.sub('\D','',node[0].split(',')[0]))
            strip = int(re.sub('\D', '', node[0].split(',')[1]))
            singleStrList = []
            for xy in self.dic_Str[(dam,strip)][8:]:
                if xy:
                    x = float(xy.split(',')[0])
                    y = float(xy.split(',')[1])
                    singleStrList.append([round(x,3), round(y,3)])
            order_2ds = self.adjust_pts_order(singleStrList)
            Str_xy_list[CC-1].append(order_2ds)

        for i in range(1, len(self.dic_CC)+1):
            obj = Str_xy_list[i-1]
            Str_xy_dic[i] = obj

        return Str_xy_dic

    # 顺时针排序
    def adjust_pts_order(self, pts_2ds):
        cen_x, cen_y = np.mean(pts_2ds, axis=0)
        d2s = []
        for i in range(len(pts_2ds)):
            o_x = pts_2ds[i][0] - cen_x
            o_y = pts_2ds[i][1] - cen_y
            atan2 = np.arctan2(o_y, o_x)
            if atan2 < 0:
                atan2 += np.pi * 2
            d2s.append([pts_2ds[i], atan2])
        d2s = sorted(d2s, key=lambda x: x[1])
        order_2ds = np.array([x[0] for x in d2s])
        return order_2ds

if __name__== "__main__":
    app = QApplication(sys.argv)
    form = myMainWindow()
    form.show()
    sys.exit(app.exec_())